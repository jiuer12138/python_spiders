# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
from openpyxl import Workbook
import urllib.request
import os

from openpyxl.reader.excel import load_workbook
from scrapy.utils.project import get_project_settings


# from KJYF_spider.main import get_data


class KjyfSpiderPipeline:
    settings = get_project_settings()

    def __init__(self):
        super(KjyfSpiderPipeline, self).__init__()
        # self.wb = Workbook()
        # self.ws = self.wb.active
        # self.ws.append(self.settings.get('CUSTOM_EXECL_HEADERS'))

    def process_item(self, item, spider):
        # 兼容文件夹不存在的情况
        execl_path = item['dir_path'].rstrip("\\")
        file_name = self.settings.get('CUSTOM_EXECL_FILE_NAME')
        file_path = execl_path + '/' + file_name + '.xlsx'
        file_header = self.settings.get('CUSTOM_EXECL_HEADERS')
        if not os.path.exists(execl_path):
            os.makedirs(execl_path)
        elif not os.path.exists(file_path):
            write_excel_xlsx_append(file_path, file_header, True)
        line = [item['cate_id'], item['p_id'], item['title'], item['price'], item['spec'],
                item['images']]
        write_excel_xlsx_append(file_path, line)
        return item


def create_excel_xlsx(path, sheet_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    workbook.save(path)


def write_excel_xlsx_append(path, value, truncate_sheet=False):
    # 如果不存在就创建该excel
    if not os.path.exists(path):
        create_excel_xlsx(path, 'Sheet')

    # value = value.values  # 将dataframe转为array
    data = load_workbook(path)
    # 取第一张表
    sheetnames = data.sheetnames
    sheet = data[sheetnames[0]]
    sheet = data.active
    if (truncate_sheet):  # truncate_sheet为True，覆盖原表中的数据
        startrows = 0
    else:
        # print(sheet.title)  # 输出表名
        startrows = sheet.max_row  # 获得行数

    for j in range(0, len(value)):
        sheet.cell(row=startrows + 1, column=j + 1, value=str(value[j]))
    data.save(path)
    print("xlsx格式表格追加写入数据成功！")


# def read_excel_xlsx(path, sheet_name):
#     workbook = load_workbook(path)
#     # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
#     sheet = workbook[sheet_name]
#     for row in sheet.rows:
#         for cell in row:
#             print(cell.value, "\t", end="")
#         print()
